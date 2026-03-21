"""
Deal Sourcing 데이터베이스 빌더
- Google Sheets에서 직접 다운로드
- New Dealsourcing (2025, 2026) + 통합(2023-2026) 탭 취합
- 2026 탭 헤더 기준으로 통합 엑셀 파일 생성
"""

import os
import ssl
import urllib.request
import subprocess
import pandas as pd
import io
import re
import json
import logging
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

ssl._create_default_https_context = ssl._create_unverified_context

# ── 로깅 설정 ──
LOG_FILE = "/Users/juyoungeun/Project/dealmap-pages/deal_db_builder.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)

SHEET_ID = "1Sg5R-kVXuAjVyfkBkBq-5S1ntdGxzsJSZIUFq1xELU8"
OUTPUT_FILE = "/Users/juyoungeun/Project/dealmap-pages/deal_database.xlsx"
OUTPUT_JSON = "/Users/juyoungeun/Project/dealmap-pages/deals.json"

# ── 취합 대상 탭 (탭 이름에 포함되는 키워드) ──
TARGET_DEALSOURCING = ["2025", "2026"]  # 2024 제외

# ── 2026 기준 통합 헤더 (출력 컬럼) ──
HEADER = [
    "No.",                           # A
    "Year",                          # B
    "Source_Tab",                     # C
    "Initial date of Review",        # D  ← 날짜 yyyy-mm-dd
    "Asset Type",                    # E
    "Existing conditions",           # F
    "Development Type",              # G
    "Development Structure",         # H
    "Project Name",                  # I
    "Comment_t",                     # J
    "지구단위계획구역",                # K
    "Location_City",                 # L
    "Location_District",             # M
    "Address",                       # N
    "Zoning District",               # O
    "Zoning District(Eng.)",         # P
    "Location Tier",                 # Q
    "Sourcing Channel",              # R
    "Asking Price",                  # S  ← 천단위 콤마, 소수점 2자리
    "Land price/Pyeong",             # T
    "FAR(%)",                        # U
    "FAR/Pyeong",                    # V
    "Land area(m2)",                 # W
    "Land area(py)",                 # X
    "Building area(m2)",             # Y
    "Building area(py)",             # Z
    "Gross floor area(m2)",          # AA
    "Gross floor area(py)",          # AB
    "# of Units(Remodeling)",        # AC
    "Price/Unit",                    # AD
    "NRA(pyeong)",                   # AE ← 여기까지 숫자 서식
    "Details",                       # AF
    "LOCATION",                      # AG
    "PRICE",                         # AH
    "ENG.",                          # AI
    "Status",                        # AJ
    "Review Stage",                  # AK
]

# S(19)~AE(31) = 인덱스 18~30 (0-based)
NUMERIC_COL_START = 18  # "Asking Price" = S열
NUMERIC_COL_END = 30    # "NRA(pyeong)"  = AE열

NUMERIC_COLS = HEADER[NUMERIC_COL_START:NUMERIC_COL_END + 1]


def extract_year_from_tab(tab_name):
    """탭 이름에서 연도 추출"""
    m = re.search(r"(20\d{2})", tab_name)
    return int(m.group(1)) if m else None


def extract_dong_bunji(project_name):
    """
    Project Name에서 동 및 번지 정보 추출
    예: '철산동 447 부지 개발사업' → '철산동 447'
        '장안동 432-3 개발사업' → '장안동 432-3'
        '남산동3가 13-21, 13-14 명동MOM하우스' → '남산동3가 13-21, 13-14'
    """
    if not project_name or pd.isna(project_name):
        return None
    s = str(project_name).strip()

    # 패턴: [도/시/구/군/읍/면 +] 동/리이름(+숫자가) + 번지
    # 예: 하남시 망월동 1146 / 경기도 용인시 기흥구 고매동 263-51
    #     하남 신장동 610 / 설성면 대죽리 1087-16 / 청파동 1가 180-24, 18
    m = re.match(
        r"((?:[가-힣]+(?:도|시|구|군|읍|면)?\s+)*"  # 선택적 도/시/구/군/읍/면 또는 약칭(하남 등)
        r"[가-힣]+(?:\d+가)?)\s+"                    # 동/리 이름 (예: 망월동, 대죽리, 남산동3가)
        r"(\d+(?:-\d+)?(?:\s*,\s*\d+(?:-\d+)?)*)",   # 번지
        s,
    )
    if m:
        dong = m.group(1)
        bunji = m.group(2)
        # "1가" 뒤에 실제 번지가 올 경우 (예: "청파동 1가 180-24")
        rest = s[m.end():].strip()
        rest_m = re.match(r"(\d+(?:-\d+)?(?:\s*,\s*\d+(?:-\d+)?)*)", rest)
        if dong.endswith("가") and rest_m:
            return f"{dong} {bunji}가 {rest_m.group(1)}"
        # "가 번지" 패턴 (예: "청파동 1" + "가 180-24")
        rest2_m = re.match(r"가\s+(\d+(?:-\d+)?(?:\s*,\s*\d+(?:-\d+)?)*)", rest)
        if rest2_m:
            return f"{dong} {bunji}가 {rest2_m.group(1)}"
        return f"{dong} {bunji}"
    return None


def parse_dealsourcing_tab(df_raw, tab_name, year):
    """
    New Dealsourcing (2025/2026) 탭 파싱
    - 헤더: Row 10-11, 데이터: Row 12~
    """
    rows = []

    header_row = None
    for i in range(min(15, len(df_raw))):
        vals = [str(v) for v in df_raw.iloc[i]]
        if "No." in vals:
            header_row = i
            break

    if header_row is None:
        log.warning(f"  [SKIP] {tab_name}: 헤더를 찾을 수 없음")
        return rows

    data_start = header_row + 2

    for i in range(data_start, len(df_raw)):
        r = df_raw.iloc[i]
        no_val = r.iloc[1]

        try:
            no_int = int(float(no_val))
        except (ValueError, TypeError):
            continue

        # 2025, 2026: Col8=Comment_t, Col9=지구단위, Col10=City, Col11=District
        row = {
            "No.": no_int,
            "Year": year,
            "Source_Tab": tab_name,
            "Initial date of Review": r.iloc[2],
            "Asset Type": r.iloc[3],
            "Existing conditions": r.iloc[4],
            "Development Type": r.iloc[5],
            "Development Structure": r.iloc[6],
            "Project Name": r.iloc[7],
            "Comment_t": r.iloc[8],
            "지구단위계획구역": r.iloc[9],
            "Location_City": r.iloc[10],
            "Location_District": r.iloc[11],
            "Address": None,
            "Zoning District": r.iloc[12],
            "Zoning District(Eng.)": r.iloc[13],
            "Location Tier": r.iloc[14],
            "Sourcing Channel": r.iloc[15],
            "Asking Price": r.iloc[16],
            "Land price/Pyeong": r.iloc[17],
            "FAR(%)": r.iloc[18],
            "FAR/Pyeong": r.iloc[19],
            "Land area(m2)": r.iloc[20],
            "Land area(py)": r.iloc[21],
            "Building area(m2)": r.iloc[22],
            "Building area(py)": r.iloc[23],
            "Gross floor area(m2)": r.iloc[24],
            "Gross floor area(py)": r.iloc[25],
            "# of Units(Remodeling)": r.iloc[26],
            "Price/Unit": r.iloc[27],
            "NRA(pyeong)": r.iloc[28],
            "Details": r.iloc[29],
            "LOCATION": r.iloc[30],
            "PRICE": r.iloc[31],
            "ENG.": r.iloc[32],
            "Status": None,
            "Review Stage": None,
        }
        rows.append(row)

    return rows


def parse_tonghap_tab(df_raw, tab_name):
    """
    통합(2023-2026) 탭 파싱
    - 헤더: Row 11-12, 데이터: Row 13~
    """
    rows = []
    data_start = 13

    for i in range(data_start, len(df_raw)):
        r = df_raw.iloc[i]
        no_val = r.iloc[1]

        try:
            no_int = int(float(no_val))
        except (ValueError, TypeError):
            continue

        year_val = r.iloc[2]
        try:
            year_int = int(float(year_val))
        except (ValueError, TypeError):
            year_int = None

        row = {
            "No.": no_int,
            "Year": year_int,
            "Source_Tab": tab_name,
            "Initial date of Review": r.iloc[3],
            "Asset Type": r.iloc[4],
            "Existing conditions": r.iloc[5],
            "Development Type": r.iloc[6],
            "Development Structure": r.iloc[7],
            "Project Name": r.iloc[8],
            "Comment_t": None,
            "지구단위계획구역": None,
            "Location_City": r.iloc[9],
            "Location_District": r.iloc[10],
            "Address": r.iloc[11],
            "Zoning District": r.iloc[12],
            "Zoning District(Eng.)": r.iloc[13],
            "Location Tier": r.iloc[14],
            "Sourcing Channel": r.iloc[15],
            "Asking Price": r.iloc[16],
            "Land price/Pyeong": r.iloc[17],
            "FAR(%)": r.iloc[18],
            "FAR/Pyeong": r.iloc[19],
            "Land area(m2)": r.iloc[20],
            "Land area(py)": r.iloc[21],
            "Building area(m2)": r.iloc[22],
            "Building area(py)": r.iloc[23],
            "Gross floor area(m2)": r.iloc[24],
            "Gross floor area(py)": r.iloc[25],
            "# of Units(Remodeling)": r.iloc[26],
            "Price/Unit": r.iloc[27],
            "NRA(pyeong)": r.iloc[28],
            "Details": r.iloc[29],
            "LOCATION": r.iloc[32],
            "PRICE": r.iloc[33],
            "ENG.": r.iloc[34],
            "Status": r.iloc[30],
            "Review Stage": r.iloc[35],
        }
        rows.append(row)

    return rows


def clean_value(v):
    """nan 제거"""
    if pd.isna(v):
        return None
    return v


def fill_address_from_project(result):
    """Address 공란을 Project Name에서 동/번지 추출하여 채움"""
    mask = result["Address"].isna() | (result["Address"] == "")
    filled = 0
    for idx in result[mask].index:
        pname = result.at[idx, "Project Name"]
        addr = extract_dong_bunji(pname)
        if addr:
            result.at[idx, "Address"] = addr
            filled += 1
    return filled


def git_push(repo_dir):
    """변경된 deals.json, deal_database.xlsx를 자동 커밋 & 푸시"""
    def run(cmd):
        return subprocess.run(
            cmd, cwd=repo_dir, capture_output=True, text=True, timeout=60
        )

    # 변경사항 확인
    st = run(["git", "status", "--porcelain", "deals.json", "deal_database.xlsx"])
    if not st.stdout.strip():
        log.info("8. Git: 변경사항 없음 — push 생략")
        return

    run(["git", "add", "deals.json", "deal_database.xlsx"])
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    result = run(["git", "commit", "-m", f"Auto-update deal database ({today})"])
    if result.returncode != 0:
        log.warning(f"8. Git commit 실패: {result.stderr.strip()}")
        return

    push = run(["git", "push"])
    if push.returncode == 0:
        log.info("8. Git push 완료")
    else:
        log.warning(f"8. Git push 실패: {push.stderr.strip()}")


def main():
    log.info("=" * 50)
    log.info(f"Deal DB 빌드 시작: {datetime.now():%Y-%m-%d %H:%M:%S}")
    log.info("=" * 50)

    log.info("1. Google Sheets 다운로드 중...")
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    response = urllib.request.urlopen(url)
    data = response.read()
    log.info(f"   다운로드 완료: {len(data):,} bytes")

    log.info("2. 엑셀 파싱 중...")
    df_dict = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None)
    log.info(f"   총 {len(df_dict)}개 탭")

    all_rows = []

    for tab_name, df_raw in df_dict.items():
        # New Dealsourcing 2025, 2026만
        if "Dealsourcing" in tab_name:
            year = extract_year_from_tab(tab_name)
            if str(year) not in TARGET_DEALSOURCING:
                log.info(f"   [SKIP] {tab_name} (대상 아님)")
                continue
            log.info(f"3. 파싱: {tab_name} (year={year})...")
            rows = parse_dealsourcing_tab(df_raw, tab_name, year)
            log.info(f"   → {len(rows)}건")
            all_rows.extend(rows)

        elif "통합" in tab_name:
            log.info(f"3. 파싱: {tab_name}...")
            rows = parse_tonghap_tab(df_raw, tab_name)
            log.info(f"   → {len(rows)}건")
            all_rows.extend(rows)

    # DataFrame 생성 및 정리
    result = pd.DataFrame(all_rows, columns=HEADER)
    result = result.map(clean_value)

    # Project Name이 없는 빈 행 제거
    before = len(result)
    result = result[result["Project Name"].notna()].reset_index(drop=True)
    removed = before - len(result)
    log.info(f"   빈 행 제거: {removed}건 (Project Name 없음)")

    # 날짜 컬럼 정리
    result["Initial date of Review"] = pd.to_datetime(
        result["Initial date of Review"], errors="coerce"
    )

    # 숫자 컬럼 정리
    for col in NUMERIC_COLS:
        result[col] = pd.to_numeric(result[col], errors="coerce")

    # Address 공란 채우기 (Project Name에서 동/번지 추출)
    filled = fill_address_from_project(result)
    log.info(f"4. Address 자동 채움: {filled}건")

    log.info(f"5. 총 {len(result)}건 취합 완료")
    log.info(f"   연도별: {result.groupby('Year').size().to_dict()}")
    log.info(f"   탭별:   {result.groupby('Source_Tab').size().to_dict()}")

    # ── 엑셀 저장 ──
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="Deal Database")
        ws = writer.sheets["Deal Database"]

        # D열 (4번째 컬럼) 날짜 서식: yyyy-mm-dd
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=4)
            cell.number_format = "YYYY-MM-DD"

        # S~AE열 (19~31번째 컬럼) 숫자 서식: 천단위 콤마 + 소수점 2자리
        for col_idx in range(NUMERIC_COL_START + 1, NUMERIC_COL_END + 2):  # 1-based
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "#,##0.00"

        # 헤더 스타일: 배경색 + 폰트
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for col_idx in range(1, len(HEADER) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 컬럼 너비: 실제 데이터 내용 기반 자동 조정
        for col_idx in range(1, len(HEADER) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row in range(2, min(ws.max_row + 1, 200)):  # 상위 200행 샘플링
                val = ws.cell(row=row, column=col_idx).value
                if val is not None:
                    # 숫자 서식 적용된 경우 포맷된 길이 추정
                    cell_len = len(str(val))
                    if isinstance(val, float):
                        cell_len = len(f"{val:,.2f}")
                    max_len = max(max_len, cell_len)
            # 최소 8, 최대 40
            width = min(max(max_len + 3, 8), 40)
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = width

    log.info(f"6. 저장 완료: {OUTPUT_FILE}")

    # ── JSON 내보내기 (웹 DealMap용) ──
    deals_json = []
    for idx, row in result.iterrows():
        # 날짜 포맷
        init_date = row["Initial date of Review"]
        if pd.notna(init_date):
            init_date = init_date.strftime("%Y-%m-%d") if hasattr(init_date, "strftime") else str(init_date)
        else:
            init_date = ""

        def str_val(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ""
            return str(v)

        def num_val(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return "-"
            return round(float(v), 2)

        deals_json.append({
            "id": str(idx),
            "Name": str_val(row["Project Name"]),
            "Address": str_val(row["Address"]),
            "Initial Date": init_date,
            "Contact Point": str_val(row["Sourcing Channel"]),
            "Asking Price": num_val(row["Asking Price"]),
            "Land Price/py": num_val(row["Land price/Pyeong"]),
            "FAR/py": num_val(row["FAR/Pyeong"]),
            "Land Area(py)": num_val(row["Land area(py)"]),
            "Deal Status": "중단",
            "Note": "",
        })

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(deals_json, f, ensure_ascii=False, indent=2)

    log.info(f"7. JSON 저장 완료: {OUTPUT_JSON} ({len(deals_json)}건)")

    # ── Git auto-push ──
    repo_dir = os.path.dirname(os.path.abspath(__file__))
    git_push(repo_dir)
    log.info("=" * 50)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.error(f"실행 실패: {e}", exc_info=True)
